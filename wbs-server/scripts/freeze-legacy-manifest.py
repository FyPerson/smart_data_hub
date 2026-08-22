# -*- coding: utf-8 -*-
"""
历史台账归档 · 口径冻结脚本

口径真相源：docs/local/历史台账归档/历史台账归档_方案_20260820_v1.0.md §2
本脚本为该口径的可执行冻结——固化此前 pandas 分析得出的切分逻辑，产出
逐 sheet 保留/剔除的 Excel 行号清单（legacy-archive-manifest.json）。
后续 Node 导入脚本（A2）不得自行解析日期做切分，只按本 manifest 选行；
双解析器（本脚本用 Python、导入脚本用 Node）互不可证，因此以本 manifest
为唯一真相源。

守卫层说明（Opus 预筛 + codex 外部审两轮吸收，口径算法本身未动）：
- 第一轮（H1/H2/M3/M4/M7/M8/L9/L10/L11）：
  行号锚定断言（H1，含 openpyxl values-only 跨解析器抽查）；
  manifest 原子落盘（H2，失败路径不触碰既有 manifest）；
  登记日期列 dtype 硬断言（M3）；
  blank 计数改用 openpyxl 全文件真实盘点（M4，read_only 模式——普通
  load_workbook 会按实际触碰单元格重新收窄 max_row，看不到部分声明
  范围内的空白尾行；pandas 自身视角对这类行天然失明）；
  manifest 补 column_count / headers[]（M7）与 environment 版本三件（M8）；
  bug_list 保留/剔除边界日期冻结断言（L9）；SOURCE_FILE 转 CLI 可选参数（L10）。
- 第二轮（H 统一异常出口 / M1 三集合契约 / M2 三点抽查 / L2）：
  main 流程加统一异常出口，覆盖此前会裸抛 traceback 的 OSError /
  zipfile.BadZipFile / openpyxl 格式异常 / KeyError / ValueError，
  改为结构化 `[FAIL] <阶段> <异常类名>: <消息>`（--debug 开关保留完整
  traceback；write_manifest 写入/替换失败单独标 manifest_write_failed）；
  每 dataset 增 blank_excel_rows 行号数组，新增 keep/drop/blank 三集合
  两两不交 + 并集覆盖 openpyxl 全文件行范围的断言（补上"blank 计数对但
  行号集合本身没被校验"的空档）；跨解析器抽查从首末两点扩到首/中/末
  三点，且数值归一吸收 pandas float 与 openpyxl int 的显示差异假红。

用法：
    "C:\\Program Files\\Python312\\python.exe" freeze-legacy-manifest.py [源xlsx路径] [--debug]
    不传路径参数时使用冻结口径记录的默认源文件。

依赖：pandas、openpyxl（仅这两个第三方库 + 标准库）。
"""

import argparse
import hashlib
import json
import os
import sys
import zipfile
from datetime import datetime, timezone

sys.stdout.reconfigure(encoding="utf-8")

import openpyxl  # noqa: E402
import pandas as pd  # noqa: E402
from openpyxl.utils.exceptions import InvalidFileException  # noqa: E402

# ---------------------------------------------------------------------------
# 常量：源文件 / 输出路径 / 冻结口径（§2 表逐字对齐，不得偏离）
# ---------------------------------------------------------------------------

SCRIPT_VERSION = "1.2"

DEFAULT_SOURCE_FILE = (
    r"E:\数据开发与治理规范手册\tmp\BMS 3.0系统 bug、维护单清单260820.xlsx"
)
OUTPUT_MANIFEST = (
    r"E:\数据开发与治理规范手册\tmp\legacy-archive-manifest.json"
)

EXPECTED_SHA256 = (
    "abe0244cade3c8fa20d006021d0f09f5a5f014e2bcd2ae1cc35bd1379319b989"
)

CUTOFF_DATE = "2026-08-01"

RESERVED_SHEET = "WpsReserved_CellImgList"

# (sheet_name, dataset_key) —— 顺序即 §2 表顺序
DATASET_SHEETS = [
    ("bug及小优化", "bug_list"),
    ("BMS维护单", "bms_maint"),
    ("财务RPA维护单", "rpa_maint"),
    ("智慧数据维护单", "smartdata_maint"),
    ("BMS数据修正单", "bms_correction"),
    ("成都贝尔SaaS平台", "chengdu_saas"),
    ("经营利润报表问题", "profit_report_issues"),
    ("OCR派工单更新记录", "ocr_dispatch_log"),
]

BUG_LIST_DATE_COLUMN = "登记日期"

# 逐 dataset 预期计数（断言用）。blank 为 M4 openpyxl 全文件真实盘点值
# （不是 pandas 视角——pandas 会静默截掉部分声明范围内的空白尾行）；
# column_count 为 pandas df.shape[1] 实测值（脚本内先跑一次取数后写死，M7）。
EXPECTED_COUNTS = {
    "bug_list": {
        "keep": 8232,
        "drop": 27,
        "blank": 0,
        "undated_kept": 9,
        "parse_failed": 0,
        "column_count": 51,
    },
    "bms_maint": {
        "keep": 378,
        "drop": 0,
        "blank": 21,
        "undated_kept": 0,
        "parse_failed": 0,
        "column_count": 19,
    },
    "rpa_maint": {
        "keep": 5,
        "drop": 0,
        "blank": 14,
        "undated_kept": 0,
        "parse_failed": 0,
        "column_count": 9,
    },
    "smartdata_maint": {
        "keep": 43,
        "drop": 0,
        "blank": 2,
        "undated_kept": 0,
        "parse_failed": 0,
        "column_count": 8,
    },
    "bms_correction": {
        "keep": 27,
        "drop": 0,
        "blank": 32,
        "undated_kept": 0,
        "parse_failed": 0,
        "column_count": 5,
    },
    "chengdu_saas": {
        "keep": 7,
        "drop": 0,
        "blank": 0,
        "undated_kept": 0,
        "parse_failed": 0,
        "column_count": 7,
    },
    "profit_report_issues": {
        "keep": 26,
        "drop": 0,
        "blank": 0,
        "undated_kept": 0,
        "parse_failed": 0,
        "column_count": 9,
    },
    "ocr_dispatch_log": {
        "keep": 15,
        "drop": 0,
        "blank": 0,
        "undated_kept": 0,
        "parse_failed": 0,
        "column_count": 11,
    },
}

EXPECTED_TOTAL_KEEP = 8733
EXPECTED_TOTAL_BLANK = 69

# L9：bug_list 保留/剔除边界日期冻结（本数据实测；换源文件时边界漂移会被拦）
EXPECTED_BUG_LIST_MAX_KEEP_DATE = pd.Timestamp("2026-07-31")
EXPECTED_BUG_LIST_MIN_DROP_DATE = pd.Timestamp("2026-08-03")

# H·统一异常出口：这些异常类型此前会以裸 Python traceback 崩出，
# 现改为结构化 [FAIL] <阶段> <异常类名>: <消息>（--debug 时仍还原完整 traceback）。
GUARDED_EXCEPTIONS = (OSError, zipfile.BadZipFile, InvalidFileException, KeyError, ValueError)


class AssertionFailure(Exception):
    """携带"期望 vs 实际"结构化差异的断言失败。"""

    def __init__(self, label, expected, actual):
        self.label = label
        self.expected = expected
        self.actual = actual
        super().__init__(f"{label}: expected={expected!r} actual={actual!r}")


def parse_args(argv=None):
    parser = argparse.ArgumentParser(description="历史台账归档 · 口径冻结脚本")
    parser.add_argument(
        "source_file",
        nargs="?",
        default=DEFAULT_SOURCE_FILE,
        help=(
            "源 xlsx 路径（默认=冻结口径记录的路径）。"
            "仅允许传入与冻结 SHA-256（见 EXPECTED_SHA256）一致的源文件路径副本——"
            "换一份内容不同的文件会在 source_sha256 断言处判红，这是设计意图而非缺陷。"
            "若确需针对新数据重新冻结口径，必须同步修改脚本内 EXPECTED_SHA256、"
            "EXPECTED_COUNTS、EXPECTED_TOTAL_*、L9 边界日期常量，不能只换文件路径了事。"
        ),
    )
    parser.add_argument(
        "--debug",
        action="store_true",
        help="出错时保留完整 Python traceback（默认只打印结构化 [FAIL] <阶段> 摘要）",
    )
    return parser.parse_args(argv)


def check_expected_counts_coverage(expected_counts, dataset_sheets):
    """L11：防新表漏配 EXPECTED_COUNTS 导致后续 KeyError——集合对拍必须先行。"""
    expected_keys = set(expected_counts.keys())
    actual_keys = {dataset_key for _, dataset_key in dataset_sheets}
    if expected_keys != actual_keys:
        raise AssertionFailure(
            "expected_counts_coverage",
            sorted(actual_keys),
            {
                "actual": sorted(expected_keys),
                "missing_in_expected": sorted(actual_keys - expected_keys),
                "extra_in_expected": sorted(expected_keys - actual_keys),
            },
        )


def check_sha256(source_file, expected_sha256):
    """断言第一条：源文件 SHA-256 必须与冻结口径记录的值一致。"""
    h = hashlib.sha256()
    with open(source_file, "rb") as f:
        while True:
            chunk = f.read(1024 * 1024)
            if not chunk:
                break
            h.update(chunk)
    actual = h.hexdigest()
    if actual != expected_sha256:
        raise AssertionFailure("source_sha256", expected_sha256, actual)
    return actual


def check_sheet_set(actual_sheet_names, dataset_sheets, reserved_sheet):
    """sheet 集合必须恰好=8 张数据 sheet + 1 张 WpsReserved 内部表，多一张/少一张都判红。"""
    expected_set = {name for name, _ in dataset_sheets} | {reserved_sheet}
    actual_set = set(actual_sheet_names)
    if actual_set != expected_set:
        missing = expected_set - actual_set
        extra = actual_set - expected_set
        raise AssertionFailure(
            "sheet_set",
            sorted(expected_set),
            {"actual": sorted(actual_set), "missing": sorted(missing), "extra": sorted(extra)},
        )


def find_blank_row_index_set(df):
    """全空行判定（pandas 视角，供 keep/drop 构建 + H1.2 区间锚定用）：
    df.dropna(how='all') 会剔除的行——既不入 keep 也不入 drop。
    注意：这不是 manifest 里 counts.blank / blank_excel_rows 的来源——那组
    字段改用 M4 的 openpyxl 全文件真实盘点（见 count_blank_rows），因为
    pandas 在这份源文件上会把部分声明范围内的空白尾行整行吞掉、连 NaN 行
    都不返回，此处的检测口径因此对这类行天然失明。"""
    non_blank_mask = ~df.isna().all(axis=1)
    return set(df.index[~non_blank_mask])


def count_blank_rows(ws):
    """M4：blank 计数改真承重。openpyxl 在 [2, ws.max_row] 区间逐行判
    "全部单元格 None/空串"，同时返回计数与行号集合（M1：manifest 需要
    落地 blank_excel_rows 数组供三集合断言使用）。ws.max_row 取自
    read_only 模式下对 sheet 声明尺寸的信任读取（普通 load_workbook 会
    按实际触碰单元格重新收窄 max_row，从而看不到这些尾部空白行；已用
    两种模式独立探测验证：read_only=True 才能复现预筛给出的 69 行）。"""
    blank_rows = []
    # 用 enumerate 显式给行号定序，不依赖 row[0].row——read_only 模式下全空
    # 的行，其首格常被 openpyxl 实体化成 EmptyCell（纯样式占位符），这个
    # 类型没有 .row/.column/.coordinate 属性，取 row[0].row 会直接 AttributeError
    # 崩溃（首跑即实锤：本文件 bms_maint 等表就存在这类首格为空的空白行）。
    for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        is_blank = True
        for cell in row:
            v = cell.value
            if v is not None and str(v).strip() != "":
                is_blank = False
                break
        if is_blank:
            blank_rows.append(row_num)
    return len(blank_rows), blank_rows


def first_named_column(columns):
    """返回首个非 'Unnamed: N' 列名的 0 基下标；理论上不应为 None。"""
    for i, c in enumerate(columns):
        if not str(c).startswith("Unnamed:"):
            return i
    return None


def normalize_spot_value(value):
    """跨解析器抽查值归一：
    - NaN/None → 空串；
    - M2：float 且 is_integer() → str(int(v))——杜绝 pandas 把整列上转型为
      float64（只因列内混了个别缺失值）后显示成 "1.0"，而 openpyxl 原样
      读回 "1" 这类纯格式化差异被误判为跨解析器不一致；
    - 其余 str() 化去首尾空白。"""
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def spot_check_row(ws, df, excel_row, col_idx, col_name):
    """H1.3：跨解析器抽查——同一 excel_row 同一列，pandas 读值 vs
    openpyxl（values-only）读值，str 化归一后比对。
    idx = excel_row - 2 是对 build_manifest 里 excel_row=idx+2 正向变换的
    反向重建；若行号计算本身跑偏（如 off-by-one 变异），反向重建出的 idx
    可能落到 df 索引范围外——这不该让脚本崩溃，而应作为一条可读的
    MISMATCH 呈现给上层断言收集，故显式兜底 KeyError/IndexError。"""
    idx = excel_row - 2
    try:
        pandas_val = normalize_spot_value(df.at[idx, col_name])
    except (KeyError, IndexError):
        pandas_val = f"<pandas idx={idx} 越界，df 有效范围 0..{len(df) - 1}>"
    openpyxl_val = normalize_spot_value(ws.cell(row=excel_row, column=col_idx + 1).value)
    return pandas_val, openpyxl_val


def process_bug_list(df, cutoff_date):
    """bug_list 专属口径：登记日期切分 + 无日期照收 + 解析失败异常计数。"""
    blank_index_set = find_blank_row_index_set(df)

    # M3：登记日期列 dtype 必须已是 datetime64——read_excel 后若仍是 object
    # （WPS 把日期列重存成文本），后面的 to_datetime 就成了对文本做二次解析
    # 而非恒等变换，parse_failed 逻辑不能再假设"极少触发"。
    date_dtype_ok = pd.api.types.is_datetime64_any_dtype(df[BUG_LIST_DATE_COLUMN])

    parsed_dates = pd.to_datetime(df[BUG_LIST_DATE_COLUMN], errors="coerce")
    cutoff_ts = pd.Timestamp(cutoff_date)

    keep_rows = []
    drop_rows = []
    undated_kept = 0
    parse_failed = 0
    max_keep_date = None
    min_drop_date = None

    for idx in df.index:
        if idx in blank_index_set:
            continue  # 全空行：不入 keep 也不入 drop

        excel_row = int(idx) + 2  # 表头占第 1 行
        date_val = parsed_dates.loc[idx]
        orig_val = df.loc[idx, BUG_LIST_DATE_COLUMN]

        if pd.isna(date_val):
            if pd.isna(orig_val):
                # NaT 且原值为空 → 无日期照收
                undated_kept += 1
                keep_rows.append(excel_row)
            else:
                # NaT 但原值非空 → "有值但解析失败"，keep 但计入异常
                parse_failed += 1
                keep_rows.append(excel_row)
        else:
            if date_val < cutoff_ts:
                keep_rows.append(excel_row)
                if max_keep_date is None or date_val > max_keep_date:
                    max_keep_date = date_val
            else:
                drop_rows.append(excel_row)
                if min_drop_date is None or date_val < min_drop_date:
                    min_drop_date = date_val

    counts = {
        "keep": len(keep_rows),
        "drop": len(drop_rows),
        "undated_kept": undated_kept,
        "parse_failed": parse_failed,
    }
    extra = {
        "df_len": len(df),
        "interior_blank_excel_rows": {int(i) + 2 for i in blank_index_set},
        "date_dtype_ok": date_dtype_ok,
        "date_bounds": {"max_keep_date": max_keep_date, "min_drop_date": min_drop_date},
    }
    return keep_rows, drop_rows, counts, extra


def process_full_keep(df):
    """非 bug_list sheet：非全空行全 keep，drop 恒空。"""
    blank_index_set = find_blank_row_index_set(df)
    keep_rows = [int(idx) + 2 for idx in df.index if idx not in blank_index_set]

    counts = {
        "keep": len(keep_rows),
        "drop": 0,
        "undated_kept": 0,
        "parse_failed": 0,
    }
    extra = {
        "df_len": len(df),
        "interior_blank_excel_rows": {int(i) + 2 for i in blank_index_set},
        "date_dtype_ok": True,  # 非 bug_list 无该口径，恒真（不参与判红）
        "date_bounds": None,
    }
    return keep_rows, [], counts, extra


def build_manifest(xl, wb_openpyxl, cutoff_date, dataset_sheets):
    """逐 sheet 跑切分逻辑，产出 datasets 清单 + 断言用 extras（extras 不写入 manifest）。"""
    datasets = []
    extras = {}
    for sheet_name, dataset_key in dataset_sheets:
        df = pd.read_excel(xl, sheet_name=sheet_name, header=0)
        columns = [str(c) for c in df.columns]
        ws = wb_openpyxl[sheet_name]

        if dataset_key == "bug_list":
            keep_rows, drop_rows, counts, extra = process_bug_list(df, cutoff_date)
        else:
            keep_rows, drop_rows, counts, extra = process_full_keep(df)

        # M4：blank 计数改真承重——openpyxl 全文件盘点覆盖 pandas 视角计数；
        # M1：同时落地行号集合供三集合契约断言使用。
        blank_count, blank_excel_rows = count_blank_rows(ws)
        counts["blank"] = blank_count

        # H1.3/M2：跨解析器抽查——首/中/末三点 keep_excel_row，取首个具名列
        first_idx = first_named_column(columns)
        spot_results = []
        if keep_rows and first_idx is not None:
            col_name = df.columns[first_idx]
            sorted_keep = sorted(keep_rows)
            mid_excel_row = sorted_keep[len(sorted_keep) // 2]
            spot_points = (
                ("first", sorted_keep[0]),
                ("mid", mid_excel_row),
                ("last", sorted_keep[-1]),
            )
            for label, excel_row in spot_points:
                pandas_val, openpyxl_val = spot_check_row(ws, df, excel_row, first_idx, col_name)
                spot_results.append((label, excel_row, pandas_val, openpyxl_val))

        datasets.append(
            {
                "dataset_key": dataset_key,
                "sheet_name": sheet_name,
                "keep_excel_rows": keep_rows,
                "drop_excel_rows": drop_rows,
                "blank_excel_rows": blank_excel_rows,  # M1
                "counts": counts,
                "column_count": len(columns),  # M7
                "headers": columns,  # M7（Unnamed 原样）
            }
        )

        extra["spot_results"] = spot_results
        extra["first_named_col_idx"] = first_idx
        extra["ws_max_row"] = ws.max_row  # M1：三集合并集对拍锚点
        extras[dataset_key] = extra

    return datasets, extras


def run_count_assertions(datasets, expected_counts, expected_total_keep, expected_total_blank):
    """逐 dataset 计数口径对拍（含 M4 真实 blank、M7 column_count、M3 dtype、parse_failed）。"""
    failures = []

    for dataset in datasets:
        key = dataset["dataset_key"]
        expected = expected_counts[key]
        actual_counts = dataset["counts"]

        for field in ("keep", "drop", "blank", "undated_kept", "parse_failed"):
            if actual_counts[field] != expected[field]:
                failures.append(
                    AssertionFailure(f"{key}.{field}", expected[field], actual_counts[field])
                )

        if dataset["column_count"] != expected["column_count"]:
            failures.append(
                AssertionFailure(
                    f"{key}.column_count", expected["column_count"], dataset["column_count"]
                )
            )

        # M1：blank_excel_rows 行号集合长度须等于 counts.blank（防"计数对但行号集合算错"）
        if len(dataset["blank_excel_rows"]) != actual_counts["blank"]:
            failures.append(
                AssertionFailure(
                    f"{key}.blank_excel_rows_length",
                    actual_counts["blank"],
                    len(dataset["blank_excel_rows"]),
                )
            )

        # parse_failed 非零即红灯（即便上面已按 !=0 抓到，这里再显式点名，
        # 确保"有值但解析失败"这类数据异常永不被总数吸收掩盖）。
        if actual_counts["parse_failed"] != 0:
            failures.append(
                AssertionFailure(f"{key}.parse_failed_nonzero", 0, actual_counts["parse_failed"])
            )

    total_keep = sum(d["counts"]["keep"] for d in datasets)
    if total_keep != expected_total_keep:
        failures.append(AssertionFailure("totals.keep", expected_total_keep, total_keep))

    total_blank = sum(d["counts"]["blank"] for d in datasets)
    if total_blank != expected_total_blank:
        failures.append(AssertionFailure("totals.blank", expected_total_blank, total_blank))

    return failures, total_keep, total_blank


def run_dtype_assertions(extras):
    """M3：登记日期列 dtype 硬断言（第二道网——parse_failed 逻辑仍保留兜底）。"""
    failures = []
    for key, extra in extras.items():
        if not extra["date_dtype_ok"]:
            failures.append(
                AssertionFailure(f"{key}.date_column_dtype_is_datetime64", True, False)
            )
    return failures


def run_anchor_assertions(datasets, extras):
    """H1：行号锚定断言，逐 dataset 三条——
    ①min(keep∪drop)==2 ②区间连续无洞（pandas 自身视角自洽性）
    ③跨解析器抽查首/中/末三点相等。"""
    failures = []
    for dataset in datasets:
        key = dataset["dataset_key"]
        keep_rows = dataset["keep_excel_rows"]
        drop_rows = dataset["drop_excel_rows"]
        ex = extras[key]
        union_rows = set(keep_rows) | set(drop_rows)

        # H1.1
        actual_min = min(union_rows) if union_rows else None
        if actual_min != 2:
            failures.append(AssertionFailure(f"{key}.anchor_min_row", 2, actual_min))

        # H1.2（pandas 自身返回帧内部自洽：不含 M4 那类被 pandas 整行吞掉的空白尾行）
        expected_range = set(range(2, 2 + ex["df_len"])) - ex["interior_blank_excel_rows"]
        if union_rows != expected_range:
            failures.append(
                AssertionFailure(
                    f"{key}.anchor_contiguous_range",
                    sorted(expected_range),
                    sorted(union_rows),
                )
            )

        # H1.3/M2
        if not ex["spot_results"]:
            failures.append(
                AssertionFailure(f"{key}.anchor_spot_check_missing", "非空 spot_results", [])
            )
        for label, excel_row, pandas_val, openpyxl_val in ex["spot_results"]:
            if pandas_val != openpyxl_val:
                failures.append(
                    AssertionFailure(
                        f"{key}.anchor_spot_check_{label}_row{excel_row}",
                        pandas_val,
                        openpyxl_val,
                    )
                )

    return failures


def run_three_set_assertions(datasets, extras):
    """M1：manifest 三集合契约——keep/drop/blank 两两不交，且并集精确覆盖
    openpyxl 全文件视角的 [2, ws.max_row] 行范围（不是 pandas 视角的
    [2, 2+len(df))——两者在有 M4 空白尾行的 dataset 上本就不同，这正是
    本条断言要补上的空档：H1.2 只证明 pandas 自身返回帧内部自洽，本条
    证明 keep+drop+blank 三者合起来才等于源文件的真实物理行范围）。"""
    failures = []
    for dataset in datasets:
        key = dataset["dataset_key"]
        keep = set(dataset["keep_excel_rows"])
        drop = set(dataset["drop_excel_rows"])
        blank = set(dataset["blank_excel_rows"])
        ex = extras[key]

        keep_drop_overlap = keep & drop
        if keep_drop_overlap:
            failures.append(
                AssertionFailure(f"{key}.keep_drop_disjoint", [], sorted(keep_drop_overlap))
            )
        keep_blank_overlap = keep & blank
        if keep_blank_overlap:
            failures.append(
                AssertionFailure(f"{key}.keep_blank_disjoint", [], sorted(keep_blank_overlap))
            )
        drop_blank_overlap = drop & blank
        if drop_blank_overlap:
            failures.append(
                AssertionFailure(f"{key}.drop_blank_disjoint", [], sorted(drop_blank_overlap))
            )

        union_all = keep | drop | blank
        expected_full_range = set(range(2, ex["ws_max_row"] + 1))
        if union_all != expected_full_range:
            failures.append(
                AssertionFailure(
                    f"{key}.three_set_union_full_range",
                    sorted(expected_full_range),
                    sorted(union_all),
                )
            )

    return failures


def run_boundary_assertions(extras, expected_max_keep_date, expected_min_drop_date):
    """L9：bug_list 保留/剔除边界日期冻结断言（换源文件时边界漂移会被拦）。"""
    failures = []
    bounds = extras.get("bug_list", {}).get("date_bounds")
    if not bounds:
        failures.append(AssertionFailure("bug_list.date_bounds_missing", "present", None))
        return failures

    if bounds["max_keep_date"] != expected_max_keep_date:
        failures.append(
            AssertionFailure(
                "bug_list.max_keep_date",
                str(expected_max_keep_date.date()),
                str(bounds["max_keep_date"]),
            )
        )
    if bounds["min_drop_date"] != expected_min_drop_date:
        failures.append(
            AssertionFailure(
                "bug_list.min_drop_date",
                str(expected_min_drop_date.date()),
                str(bounds["min_drop_date"]),
            )
        )
    return failures


def write_manifest(output_path, manifest):
    """H2：原子落盘——写临时文件后 os.replace 原子替换。失败路径永不调用本函数，
    因此不存在"删已有 manifest"的场景，也就不需要 cleanup 逻辑。
    注意：os.replace 保证的是进程可见原子性（读者永远只会看到"旧文件完整"或
    "新文件完整"两种状态之一，不会看到半写状态），不承诺断电持久性——本脚本是
    admin 本机手动运行的单机工具，断电/掉盘不在其威胁模型内，不做 fsync 级别
    的持久化保证。"""
    tmp_path = output_path + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)
    os.replace(tmp_path, output_path)


def main():
    args = parse_args()
    source_file = args.source_file

    print("=" * 70)
    print("历史台账归档 · 口径冻结脚本")
    print(f"源文件: {source_file}")
    print("=" * 70)

    current_phase = "config_self_check"
    try:
        # --------------------------------------------------------------
        # L11：config 自检——EXPECTED_COUNTS 覆盖面必须与 DATASET_SHEETS 一致，
        # 防止新增/改名 sheet 时漏配导致后面 KeyError 崩溃而非可读断言失败。
        # --------------------------------------------------------------
        try:
            check_expected_counts_coverage(EXPECTED_COUNTS, DATASET_SHEETS)
        except AssertionFailure as e:
            print("[FAIL] expected_counts_coverage")
            print(f"  期望: {e.expected}")
            print(f"  实际: {e.actual}")
            sys.exit(1)
        print("[PASS] expected_counts_coverage")

        # --------------------------------------------------------------
        # 断言第一条：源文件 SHA-256（L10：FileNotFoundError 结构化捕获）
        # --------------------------------------------------------------
        current_phase = "source_sha256_check"
        try:
            source_sha256 = check_sha256(source_file, EXPECTED_SHA256)
        except FileNotFoundError as e:
            print("[FAIL] source_file_not_found")
            print(f"  期望路径: {source_file}")
            print(f"  实际: 文件不存在（{e}）")
            sys.exit(1)
        except AssertionFailure as e:
            print("[FAIL] source_sha256")
            print(f"  期望: {e.expected}")
            print(f"  实际: {e.actual}")
            sys.exit(1)
        print(f"[PASS] source_sha256 = {source_sha256}")

        # --------------------------------------------------------------
        # L11：全脚本复用单个 pd.ExcelFile 对象（with 管理生命周期，显式 close）
        # --------------------------------------------------------------
        current_phase = "open_source_workbook"
        wb_openpyxl = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
        try:
            with pd.ExcelFile(source_file, engine="openpyxl") as xl:
                # ----------------------------------------------------
                # sheet 集合断言
                # ----------------------------------------------------
                current_phase = "sheet_set_check"
                try:
                    check_sheet_set(xl.sheet_names, DATASET_SHEETS, RESERVED_SHEET)
                except AssertionFailure as e:
                    print("[FAIL] sheet_set")
                    print(f"  期望: {e.expected}")
                    print(f"  实际: {e.actual}")
                    sys.exit(1)
                print(
                    f"[PASS] sheet_set = {len(DATASET_SHEETS)} 张数据 sheet "
                    f"+ 1 张 {RESERVED_SHEET}（忽略）"
                )

                # ----------------------------------------------------
                # 逐 sheet 切分 + 全部断言
                # ----------------------------------------------------
                current_phase = "build_manifest"
                datasets, extras = build_manifest(xl, wb_openpyxl, CUTOFF_DATE, DATASET_SHEETS)
        finally:
            wb_openpyxl.close()

        current_phase = "run_assertions"
        failures = []

        count_failures, total_keep, total_blank = run_count_assertions(
            datasets, EXPECTED_COUNTS, EXPECTED_TOTAL_KEEP, EXPECTED_TOTAL_BLANK
        )
        failures.extend(count_failures)
        failures.extend(run_dtype_assertions(extras))
        failures.extend(run_anchor_assertions(datasets, extras))
        failures.extend(run_three_set_assertions(datasets, extras))
        failures.extend(
            run_boundary_assertions(
                extras, EXPECTED_BUG_LIST_MAX_KEEP_DATE, EXPECTED_BUG_LIST_MIN_DROP_DATE
            )
        )

        if failures:
            print()
            print("[FAIL] 以下断言未通过：")
            for f in failures:
                print(f"  - {f.label}: 期望={f.expected!r} 实际={f.actual!r}")
            sys.exit(1)

        print()
        print("逐 sheet PASS 摘要：")
        for dataset in datasets:
            c = dataset["counts"]
            key = dataset["dataset_key"]
            print(
                f"  [PASS] {key:22s} ({dataset['sheet_name']}) "
                f"keep={c['keep']:5d} drop={c['drop']:3d} blank={c['blank']:3d} "
                f"undated_kept={c['undated_kept']:2d} parse_failed={c['parse_failed']} "
                f"cols={dataset['column_count']:2d}"
            )
            ex = extras[key]
            spot_desc = ", ".join(
                f"{label}(r{row})={'OK' if pv == ov else 'MISMATCH'}"
                for label, row, pv, ov in ex["spot_results"]
            )
            print(
                f"    [PASS] {key}.anchor min_row=2 "
                f"contiguous_range=[2,{2 + ex['df_len']}) spot_check(first/mid/last): {spot_desc}"
            )
            print(
                f"    [PASS] {key}.three_set keep∪drop∪blank == "
                f"[2,{ex['ws_max_row'] + 1}) (ws.max_row={ex['ws_max_row']}) "
                f"— 两两不交 + 并集覆盖全文件行范围"
            )

        print(f"  [PASS] totals.keep  = {total_keep} (期望 {EXPECTED_TOTAL_KEEP})")
        print(f"  [PASS] totals.blank = {total_blank} (期望 {EXPECTED_TOTAL_BLANK})")

        bounds = extras["bug_list"]["date_bounds"]
        print(
            f"  [PASS] bug_list.date_boundaries "
            f"max_keep_date={bounds['max_keep_date'].date()} "
            f"min_drop_date={bounds['min_drop_date'].date()}"
        )

        # --------------------------------------------------------------
        # 全过 → 写 manifest（M1/M7/M8 字段随 datasets/environment 一并写出）
        # --------------------------------------------------------------
        current_phase = "write_manifest"
        manifest = {
            "generated_at": datetime.now(timezone.utc).astimezone().isoformat(),
            "script_version": SCRIPT_VERSION,
            "source_file": os.path.basename(source_file),
            "source_sha256": source_sha256,
            "cutoff": CUTOFF_DATE,
            "environment": {
                "python": ".".join(map(str, sys.version_info[:3])),
                "pandas": pd.__version__,
                "openpyxl": openpyxl.__version__,
            },
            "datasets": datasets,
            "totals": {"keep": total_keep, "blank": total_blank},
        }
        try:
            write_manifest(OUTPUT_MANIFEST, manifest)
        except OSError as e:
            if args.debug:
                raise
            print(f"[FAIL] manifest_write_failed {type(e).__name__}: {e}")
            sys.exit(1)

        print()
        print(f"[OK] manifest 已写出: {OUTPUT_MANIFEST}")
        sys.exit(0)

    except SystemExit:
        raise
    except GUARDED_EXCEPTIONS as e:
        # H·统一异常出口：此前这些异常类型会以裸 traceback 崩出（如源文件损坏
        # 触发 zipfile.BadZipFile/InvalidFileException、列名缺失触发 KeyError、
        # 日期/数值解析异常触发 ValueError、磁盘/权限问题触发 OSError）。
        # --debug 时放弃结构化摘要、原样重新抛出以保留完整 traceback。
        if args.debug:
            raise
        print(f"[FAIL] {current_phase} {type(e).__name__}: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
