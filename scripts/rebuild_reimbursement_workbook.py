from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.workbook.defined_name import DefinedName


SOURCE_PATH = Path(
    r"E:\数据开发与治理规范手册\docs\local\数仓开发\_临时需求\报销统计表.xlsx"
)
TARGET_PATH = Path(
    r"E:\数据开发与治理规范手册\docs\local\数仓开发\_临时需求\报销统计表_重建版.xlsx"
)

SUMMARY_SHEET = "\u62a5\u9500\u7edf\u8ba1\u8868"
CONFIG_SHEET = "\u6c47\u603b\u914d\u7f6e"
TEMPLATE_SHEET = "\u4eba\u5458\u6a21\u677f"

SUMMARY_HEADERS = [
    "\u8239\u540d",
    "\u8239\u53f7",
    "\u76d1\u88c5\u4eba\u5458",
    "\u76d1\u88c5\u5929\u6570",
    "\u62a5\u9500\u91d1\u989d",
    "\u8865\u8d34\u91d1\u989d",
    "\u8239\u4e1c\u540d\u79f0",
    "\u76d1\u88c5\u8d39(USD)",
    "\u6c47\u7387",
    "\u4eba\u6c11\u5e01\u91d1\u989d",
    "\u5229\u6da6",
]

PERSON_HEADERS = [
    "\u8239\u540d",
    "\u8239\u53f7",
    "\u76d1\u88c5\u5929\u6570",
    "\u62a5\u9500\u91d1\u989d",
    "\u8865\u8d34\u91d1\u989d",
]

SHEET_NAME_NAME = "person_sheet_names"
MAX_DATA_ROWS = 500
MAX_PERSON_ROWS = 200


def copy_sheet_values_and_widths(src_ws, dst_ws):
    for row in src_ws.iter_rows():
        for cell in row:
            dst_ws[cell.coordinate].value = cell.value

    for key, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[key].width = dim.width


def apply_summary_layout(ws):
    for idx, header in enumerate(SUMMARY_HEADERS, start=1):
        ws.cell(1, idx).value = header

    widths = [12, 15, 14, 12, 14, 14, 14, 14, 10, 14, 12]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    header_fill = PatternFill(fill_type="solid", fgColor="D9E2F3")
    money_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    for cell in ws[1]:
        cell.fill = header_fill
    for col in ["E", "F", "H", "I", "J", "K"]:
        ws[f"{col}1"].fill = money_fill

    ws.freeze_panes = "A2"


def apply_person_layout(ws):
    for idx, header in enumerate(PERSON_HEADERS, start=1):
        ws.cell(1, idx).value = header

    widths = [12, 15, 12, 14, 14]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAD3")
    for cell in ws[1]:
        cell.fill = header_fill

    ws.freeze_panes = "A2"


def build_summary_formula() -> str:
    return (
        "=LET("
        "sheets,FILTER("
        f"{CONFIG_SHEET}!A2:A200,"
        f"{CONFIG_SHEET}!A2:A200<>\"\""
        "),"
        "rows,IFERROR("
        "DROP("
        "REDUCE(\"\",sheets,LAMBDA(acc,s,"
        "LET("
        f"data,INDIRECT(\"'\"&s&\"'!A2:E{MAX_PERSON_ROWS}\"),"
        "keep,BYROW(data,LAMBDA(r,COUNTA(r)>0)),"
        "n,SUM(--keep),"
        "IF(n=0,acc,"
        "LET("
        "filtered,FILTER(data,keep),"
        "names,IF(SEQUENCE(ROWS(filtered),1,1,1),s),"
        "block,HSTACK("
        "CHOOSECOLS(filtered,1),"
        "CHOOSECOLS(filtered,2),"
        "names,"
        "CHOOSECOLS(filtered,3),"
        "CHOOSECOLS(filtered,4),"
        "CHOOSECOLS(filtered,5)"
        "),"
        "VSTACK(acc,block)"
        ")"
        ")"
        ")"
        "))),1),\"\""
        "),"
        "rows"
        ")"
    )


def main() -> None:
    src = load_workbook(SOURCE_PATH, data_only=False)
    person_names = src.sheetnames[1:16]

    wb = Workbook()
    summary = wb.active
    summary.title = SUMMARY_SHEET
    apply_summary_layout(summary)

    for person_name in person_names:
        src_ws = src[person_name]
        dst_ws = wb.create_sheet(person_name)
        apply_person_layout(dst_ws)
        for row in range(2, min(src_ws.max_row, 200) + 1):
            for col in range(1, 6):
                dst_ws.cell(row, col).value = src_ws.cell(row, col).value

    template = wb.create_sheet(TEMPLATE_SHEET)
    apply_person_layout(template)
    template["G1"] = "\u8bf4\u660e"
    template["G2"] = (
        "\u590d\u5236\u672c\u9875\u5e76\u628asheet\u540d\u6539\u6210\u4eba\u5458\u59d3\u540d\uff0c"
        "\u586bA:E\u540e\u603b\u8868\u4f1a\u81ea\u52a8\u6c47\u603b\u3002"
    )

    config = wb.create_sheet(CONFIG_SHEET)
    config.sheet_state = "hidden"
    config["A1"] = "\u4eba\u5458Sheet\u540d\u5355"
    config["A2"] = (
        f'=LET(x,{SHEET_NAME_NAME},'
        f'FILTER(x,(x<>"{SUMMARY_SHEET}")*(x<>"{CONFIG_SHEET}")*(x<>"{TEMPLATE_SHEET}")))'
    )

    wb.defined_names.add(
        DefinedName(
            SHEET_NAME_NAME,
            attr_text='=REPLACE(GET.WORKBOOK(1),1,FIND("]",GET.WORKBOOK(1)),"")',
        )
    )

    summary["A2"] = build_summary_formula()
    for row in range(2, MAX_DATA_ROWS + 1):
        summary[f"J{row}"] = f'=IF(OR(H{row}="",I{row}=""),"",H{row}*I{row})'

    wb.save(TARGET_PATH)
    print(TARGET_PATH)


if __name__ == "__main__":
    main()
