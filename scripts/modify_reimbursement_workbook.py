from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.workbook.properties import CalcProperties


WORKBOOK_PATH = Path(
    r"E:\数据开发与治理规范手册\docs\local\数仓开发\_临时需求\报销统计表.xlsx"
)
SUMMARY_SHEET = "报销统计表"
CONFIG_SHEET = "汇总配置"
TEMPLATE_SHEET = "人员模板"
PERSON_SHEET_HEADERS = ["船名", "船号", "监装天数", "报销金额", "补贴金额"]
SUMMARY_HEADERS = [
    "船名",
    "船号",
    "监装人员",
    "监装天数",
    "报销金额",
    "补贴金额",
    "船东名称",
    "监装费(USD)",
    "汇率",
    "人民币金额",
    "利润",
]
SHEET_NAME_NAME = "person_sheet_names"
MAX_CLEAR_ROWS = 1000
MAX_SUMMARY_DATA_ROWS = 500
MAX_PERSON_SHEET_ROWS = 200


def clear_cells(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row, col).value = None


def ensure_headers(ws, headers) -> None:
    for idx, header in enumerate(headers, start=1):
        ws.cell(1, idx).value = header


def ensure_config_sheet(wb):
    if CONFIG_SHEET in wb.sheetnames:
        ws = wb[CONFIG_SHEET]
        clear_cells(ws, 1, MAX_CLEAR_ROWS, 1, 5)
    else:
        ws = wb.create_sheet(CONFIG_SHEET)

    ws["A1"] = "人员Sheet名单"
    ws["A2"] = (
        f'=LET(x,{SHEET_NAME_NAME},'
        f'FILTER(x,(x<>"{SUMMARY_SHEET}")*(x<>"{CONFIG_SHEET}")*(x<>"{TEMPLATE_SHEET}")))'
    )
    ws.sheet_state = "hidden"
    return ws


def replace_defined_name(wb) -> None:
    names = wb.defined_names
    if SHEET_NAME_NAME in names:
        del names[SHEET_NAME_NAME]

    names.add(
        DefinedName(
            SHEET_NAME_NAME,
            attr_text='=REPLACE(GET.WORKBOOK(1),1,FIND("]",GET.WORKBOOK(1)),"")',
        )
    )


def ensure_template_sheet(wb):
    if TEMPLATE_SHEET in wb.sheetnames:
        ws = wb[TEMPLATE_SHEET]
        clear_cells(ws, 1, MAX_CLEAR_ROWS, 1, 10)
    else:
        sample_person_sheet = wb[wb.sheetnames[1]]
        ws = wb.copy_worksheet(sample_person_sheet)
        ws.title = TEMPLATE_SHEET

    ensure_headers(ws, PERSON_SHEET_HEADERS)
    clear_cells(ws, 2, MAX_CLEAR_ROWS, 1, 5)
    ws["G1"] = "说明"
    ws["G2"] = "复制本页并把sheet名改成人员姓名，填A:E后总表会自动汇总。"
    return ws


def build_summary_formula() -> str:
    return (
        f'=LET('
        f'sheets,FILTER({CONFIG_SHEET}!A2:A200,{CONFIG_SHEET}!A2:A200<>""),'
        f'rows,IFERROR('
        f'DROP('
        f'REDUCE("",sheets,LAMBDA(acc,s,'
        f'LET('
        f'data,INDIRECT("\'"&s&"\'!A2:E{MAX_PERSON_SHEET_ROWS}"),'
        f'keep,BYROW(data,LAMBDA(r,COUNTA(r)>0)),'
        f'n,SUM(--keep),'
        f'IF(n=0,acc,'
        f'LET('
        f'filtered,FILTER(data,keep),'
        f'block,HSTACK('
        f'CHOOSECOLS(filtered,1),'
        f'CHOOSECOLS(filtered,2),'
        f'MAKEARRAY(ROWS(filtered),1,LAMBDA(r,c,s)),'
        f'CHOOSECOLS(filtered,3),'
        f'CHOOSECOLS(filtered,4),'
        f'CHOOSECOLS(filtered,5)'
        f'),'
        f'VSTACK(acc,block)'
        f')'
        f')'
        f')'
        f')),1),'
        f'""'
        f'),'
        f'rows'
        f')'
    )


def update_summary_sheet(wb) -> None:
    ws = wb[SUMMARY_SHEET]
    ensure_headers(ws, SUMMARY_HEADERS)
    clear_cells(ws, 2, MAX_CLEAR_ROWS, 1, len(SUMMARY_HEADERS))
    ws["A2"] = build_summary_formula()

    for row in range(2, MAX_SUMMARY_DATA_ROWS + 1):
        ws.cell(row, 10).value = f'=IF(OR(H{row}="",I{row}=""),"",H{row}*I{row})'


def main() -> None:
    wb = load_workbook(WORKBOOK_PATH)
    wb.calculation = CalcProperties(calcMode="auto", forceFullCalc=True, fullCalcOnLoad=True)

    replace_defined_name(wb)
    ensure_config_sheet(wb)
    ensure_template_sheet(wb)
    update_summary_sheet(wb)

    wb.save(WORKBOOK_PATH)
    print(WORKBOOK_PATH)


if __name__ == "__main__":
    main()
