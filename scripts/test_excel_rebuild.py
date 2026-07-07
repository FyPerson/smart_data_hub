from pathlib import Path

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName


def main() -> None:
    p = Path(r"E:\数据开发与治理规范手册\tmp\pq_test_rebuild.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "鎶ラ攢缁熻琛?
    ws.append(["鑸瑰悕", "鑸瑰彿", "鐩戣浜哄憳", "鐩戣澶╂暟", "鎶ラ攢閲戦", "琛ヨ创閲戦"])

    person = wb.create_sheet("寮犱笁")
    person.append(["鑸瑰悕", "鑸瑰彿", "鐩戣澶╂暟", "鎶ラ攢閲戦", "琛ヨ创閲戦"])

    template = wb.create_sheet("浜哄憳妯℃澘")
    template.append(["鑸瑰悕", "鑸瑰彿", "鐩戣澶╂暟", "鎶ラ攢閲戦", "琛ヨ创閲戦"])

    cfg = wb.create_sheet("姹囨€婚厤缃?)
    cfg.sheet_state = "hidden"
    cfg["A1"] = "浜哄憳Sheet鍚嶅崟"
    cfg["A2"] = '=LET(x,person_sheet_names,FILTER(x,(x<>"鎶ラ攢缁熻琛?)*(x<>"姹囨€婚厤缃?)*(x<>"浜哄憳妯℃澘")))'

    wb.defined_names.add(
        DefinedName(
            "person_sheet_names",
            attr_text='=REPLACE(GET.WORKBOOK(1),1,FIND("]",GET.WORKBOOK(1)),"")',
        )
    )

    ws["A2"] = (
        '=LET('
        'sheets,FILTER(姹囨€婚厤缃?A2:A200,姹囨€婚厤缃?A2:A200<>""),'
        'rows,IFERROR('
        'DROP('
        'REDUCE("",sheets,LAMBDA(acc,s,'
        'LET('
        'data,INDIRECT("\'"&s&"\'!A2:E200"),'
        'keep,BYROW(data,LAMBDA(r,COUNTA(r)>0)),'
        'n,SUM(--keep),'
        'IF(n=0,acc,'
        'LET('
        'filtered,FILTER(data,keep),'
        'names,IF(SEQUENCE(ROWS(filtered),1,1,1),s),'
        'block,HSTACK('
        'CHOOSECOLS(filtered,1),'
        'CHOOSECOLS(filtered,2),'
        'names,'
        'CHOOSECOLS(filtered,3),'
        'CHOOSECOLS(filtered,4),'
        'CHOOSECOLS(filtered,5)'
        '),'
        'VSTACK(acc,block)'
        ')'
        ')'
        ')'
        '))),1),"")'
        ',rows)'
    )

    wb.save(p)
    print(p)


if __name__ == "__main__":
    main()

